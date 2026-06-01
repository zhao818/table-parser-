"""
Comprehensive test suite for the table parsing engine.

Covers:
- Unit tests for Excel generation (merge cells, styles, edge cases)
- Unit tests for JSON validation
- Unit tests for LLM JSON extraction
- Integration tests with mock LLM service
- API endpoint tests with FastAPI TestClient

Run: pytest tests/ -v --cov=app --cov-report=term-missing
"""

from __future__ import annotations

import asyncio
import io
import struct
import sys
import zlib
from pathlib import Path

import pytest
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from app.excel_generator import (
    generate_excel,
    generate_excel_multi_sheet,
    generate_excel_simple,
    validate_table_json,
)
from app.llm_service import LLMService, _extract_json
from app.models import CellData, TableData


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _load_workbook_from_bytes(data: io.BytesIO):
    data.seek(0)
    return load_workbook(data)


def _get_cell_style(cell):
    return {
        "bold": cell.font.bold,
        "horizontal": cell.alignment.horizontal,
        "vertical": cell.alignment.vertical,
        "border_left": cell.border.left.style,
        "border_right": cell.border.right.style,
        "border_top": cell.border.top.style,
        "border_bottom": cell.border.bottom.style,
    }


def _make_simple_table() -> TableData:
    return TableData(
        title="测试表格",
        rows=[
            [CellData(text="姓名"), CellData(text="年龄")],
            [CellData(text="张三"), CellData(text="30")],
            [CellData(text="李四"), CellData(text="25")],
        ],
    )


def _make_colspan_table() -> TableData:
    return TableData(
        title="跨列表格",
        rows=[
            [CellData(text="跨列标题", colspan=3)],
            [CellData(text="A"), CellData(text="B"), CellData(text="C")],
            [CellData(text="1"), CellData(text="2"), CellData(text="3")],
        ],
    )


def _make_rowspan_table() -> TableData:
    return TableData(
        title="跨行表格",
        rows=[
            [CellData(text="跨行", rowspan=2), CellData(text="列1")],
            [CellData(text="列2")],
        ],
    )


def _make_complex_table() -> TableData:
    return TableData(
        title="复杂表格",
        rows=[
            [CellData(text="A1", colspan=2), CellData(text="B1")],
            [CellData(text="A2", rowspan=2), CellData(text="B2"), CellData(text="C2")],
            [CellData(text="B3"), CellData(text="C3")],
        ],
    )


# ============================================================================
# generate_excel
# ============================================================================

class TestGenerateExcelFromJson:
    """Unit tests for the main Excel generation function."""

    def test_simple_table_row_count(self):
        buf = generate_excel(_make_simple_table(), first_row_as_header=False)
        wb = _load_workbook_from_bytes(buf)
        assert wb.active.max_row == 4  # 1 title + 3 data

    def test_colspan_merges_cells(self):
        buf = generate_excel(_make_colspan_table(), first_row_as_header=False)
        wb = _load_workbook_from_bytes(buf)
        ranges = [str(r) for r in wb.active.merged_cells.ranges]
        assert len(ranges) >= 2  # title merge + colspan merge

    def test_rowspan_merges_cells(self):
        buf = generate_excel(_make_rowspan_table(), first_row_as_header=False)
        wb = _load_workbook_from_bytes(buf)
        ranges = [str(r) for r in wb.active.merged_cells.ranges]
        assert len(ranges) >= 2

    def test_all_cells_have_borders(self):
        buf = generate_excel(_make_simple_table(), first_row_as_header=False)
        wb = _load_workbook_from_bytes(buf)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                if cell.value is not None:
                    s = _get_cell_style(cell)
                    assert s["border_left"] == "thin"
                    assert s["border_right"] == "thin"
                    assert s["border_top"] == "thin"
                    assert s["border_bottom"] == "thin"

    def test_header_row_is_bold(self):
        buf = generate_excel(_make_simple_table(), first_row_as_header=True)
        wb = _load_workbook_from_bytes(buf)
        assert wb.active.cell(row=2, column=1).font.bold is True

    def test_title_is_bold_and_large(self):
        buf = generate_excel(_make_simple_table())
        wb = _load_workbook_from_bytes(buf)
        tc = wb.active.cell(row=1, column=1)
        assert tc.font.bold is True
        assert tc.font.size == 14

    def test_text_content_preserved(self):
        buf = generate_excel(_make_simple_table(), first_row_as_header=False)
        ws = _load_workbook_from_bytes(buf).active
        assert ws.cell(row=2, column=1).value == "姓名"
        assert ws.cell(row=3, column=1).value == "张三"

    def test_custom_sheet_name(self):
        buf = generate_excel(_make_simple_table(), sheet_name="CustomName")
        assert _load_workbook_from_bytes(buf).active.title == "CustomName"

    def test_override_title(self):
        buf = generate_excel(_make_simple_table(), title="覆盖标题")
        assert _load_workbook_from_bytes(buf).active.cell(row=1, column=1).value == "覆盖标题"

    def test_empty_title_skips_title_row(self):
        table = TableData(title="", rows=[[CellData(text="A")]])
        buf = generate_excel(table, first_row_as_header=False)
        assert _load_workbook_from_bytes(buf).active.max_row == 1


# ============================================================================
# generate_excel_simple
# ============================================================================

class TestGenerateExcelSimple:
    """Tests for the simplified header+data interface."""

    def test_basic_generation(self):
        buf = generate_excel_simple(
            headers=["姓名", "年龄", "城市"],
            data_rows=[["张三", "30", "北京"], ["李四", "25", "上海"]],
            title="用户列表",
        )
        ws = _load_workbook_from_bytes(buf).active
        assert ws.cell(row=1, column=1).value == "用户列表"
        assert ws.cell(row=2, column=1).value == "姓名"
        assert ws.cell(row=2, column=1).font.bold is True
        assert ws.cell(row=3, column=1).value == "张三"

    def test_no_title(self):
        buf = generate_excel_simple(headers=["A"], data_rows=[])
        ws = _load_workbook_from_bytes(buf).active
        assert ws.max_row == 1
        assert ws.cell(row=1, column=1).value == "A"

    def test_empty_data(self):
        buf = generate_excel_simple(headers=["H1", "H2", "H3"], data_rows=[])
        ws = _load_workbook_from_bytes(buf).active
        assert ws.max_row == 1
        assert ws.max_column == 3


# ============================================================================
# generate_excel_multi_sheet
# ============================================================================

class TestGenerateExcelMultiSheet:
    """Tests for multi-sheet workbook generation."""

    def test_multiple_sheets(self):
        sheets = {"用户表": _make_simple_table(), "物料表": _make_colspan_table()}
        buf = generate_excel_multi_sheet(sheets, title="多Sheet工作簿")
        wb = _load_workbook_from_bytes(buf)
        assert len(wb.sheetnames) == 2
        assert "用户表" in wb.sheetnames

    def test_first_sheet_has_title(self):
        sheets = {"SheetA": _make_simple_table()}
        buf = generate_excel_multi_sheet(sheets, title="全局标题")
        assert _load_workbook_from_bytes(buf)["SheetA"].cell(row=1, column=1).value == "全局标题"


# ============================================================================
# validate_table_json
# ============================================================================

class TestValidateTableJson:
    """Tests for JSON validation."""

    def test_valid_json(self):
        data = {"title": "", "rows": [[{"text": "A", "rowspan": 1, "colspan": 1}]]}
        result = validate_table_json(data)
        assert isinstance(result, TableData)
        assert result.rows[0][0].text == "A"

    def test_missing_rows(self):
        with pytest.raises(ValueError, match="must contain 'rows'"):
            validate_table_json({})

    def test_rows_not_list(self):
        with pytest.raises(ValueError, match="must be a list"):
            validate_table_json({"rows": "not a list"})

    def test_not_a_dict(self):
        with pytest.raises(ValueError, match="must be a dict"):
            validate_table_json("not a dict")

    def test_default_span(self):
        result = validate_table_json({"title": "", "rows": [[{"text": "X"}]]})
        assert result.rows[0][0].rowspan == 1
        assert result.rows[0][0].colspan == 1


# ============================================================================
# LLM JSON extraction
# ============================================================================

class TestLLMJsonExtraction:
    """Tests for JSON extraction from LLM text."""

    def test_direct_json(self):
        text = '{"title":"T","rows":[[{"text":"A","rowspan":1,"colspan":1}]]}'
        assert _extract_json(text)["title"] == "T"

    def test_markdown_json_fence(self):
        text = '```json\n{"title":"X","rows":[[{"text":"B"}]]}\n```'
        assert _extract_json(text)["title"] == "X"

    def test_markdown_code_fence(self):
        text = '```\n{"title":"Y","rows":[[{"text":"C"}]]}\n```'
        assert _extract_json(text)["title"] == "Y"

    def test_json_within_text(self):
        text = 'Preamble... {"title":"Z","rows":[[{"text":"D"}]]} trailing'
        assert _extract_json(text)["title"] == "Z"

    def test_invalid_json_raises(self):
        with pytest.raises(ValueError, match="Failed to extract"):
            _extract_json("Not JSON")


# ============================================================================
# LLMService mock
# ============================================================================

class TestLLMServiceMock:
    """Integration tests using mock LLM service."""

    def test_mock_returns_table_data(self):
        svc = LLMService(use_mock=True)
        table = asyncio.run(svc.analyze(b"fake bytes", "PNG"))
        assert isinstance(table, TableData)
        assert len(table.rows) > 0

    def test_mock_has_merged_cells(self):
        svc = LLMService(use_mock=True)
        table = svc.get_mock_table()
        has_merge = any(
            cell.rowspan > 1 or cell.colspan > 1
            for row in table.rows for cell in row
        )
        assert has_merge

    def test_factory_function(self):
        from app.llm_service import create_llm_service
        assert create_llm_service(use_mock=True)._mock is True


# ============================================================================
# Edge cases
# ============================================================================

class TestEdgeCases:
    """Boundary and robustness tests."""

    def test_single_cell(self):
        table = TableData(title="单格", rows=[[CellData(text="唯一单元格")]])
        buf = generate_excel(table, first_row_as_header=False)
        assert _load_workbook_from_bytes(buf).active.cell(row=2, column=1).value == "唯一单元格"

    def test_empty_text_cells(self):
        table = TableData(title="", rows=[[CellData(text=""), CellData(text="有内容")]])
        buf = generate_excel(table, first_row_as_header=False)
        assert _load_workbook_from_bytes(buf).active.cell(row=1, column=1).value in ("", None)

    def test_long_text(self):
        long_text = "这是一个非常长的单元格内容" * 20
        table = TableData(rows=[[CellData(text=long_text)]])
        buf = generate_excel(table, first_row_as_header=False)
        assert _load_workbook_from_bytes(buf).active.cell(row=1, column=1).value == long_text

    def test_special_characters(self):
        special = '☑ 合格  □ 不合格\n备注: <test> & "quote"'
        table = TableData(rows=[[CellData(text=special)]])
        buf = generate_excel(table, first_row_as_header=False)
        assert "☑" in _load_workbook_from_bytes(buf).active.cell(row=1, column=1).value

    def test_max_rowspan_colspan(self):
        table = TableData(rows=[[CellData(text="大合并", rowspan=5, colspan=3)]])
        buf = generate_excel(table, first_row_as_header=False)
        ranges = [str(r) for r in _load_workbook_from_bytes(buf).active.merged_cells.ranges]
        assert len(ranges) >= 1

    def test_buffer_rewound(self):
        buf = generate_excel(_make_simple_table(), first_row_as_header=False)
        assert buf.tell() == 0
        assert _load_workbook_from_bytes(buf) is not None


# ============================================================================
# API endpoints
# ============================================================================

class TestAPIEndpoints:
    """End-to-end tests for API routes."""

    @pytest.fixture(autouse=True)
    def _setup(self):
        from app.main import create_app
        from fastapi.testclient import TestClient
        self.app = create_app()
        self.app.state.llm_service = LLMService(use_mock=True)
        self.client = TestClient(self.app)

    def test_health_check(self):
        resp = self.client.get("/api/health")
        assert resp.status_code == 200
        data = resp.json()
        assert data["success"] is True
        assert data["data"]["mock_mode"] is True

    def test_upload_no_file_returns_422(self):
        assert self.client.post("/api/upload").status_code == 422

    def test_upload_image_returns_excel(self):
        def make_png():
            def chunk(ctype, data):
                c = ctype + data
                return struct.pack(">I", len(data)) + c + struct.pack(">I", zlib.crc32(c) & 0xFFFFFFFF)
            ihdr = struct.pack(">IIBBBBB", 1, 1, 8, 2, 0, 0, 0)
            return b"\x89PNG\r\n\x1a\n" + chunk(b"IHDR", ihdr) + chunk(b"IDAT", zlib.compress(b"\x00\x00\x00")) + chunk(b"IEND", b"")

        resp = self.client.post(
            "/api/upload?use_mock=true",
            files={"file": ("test.png", make_png(), "image/png")},
        )
        assert resp.status_code == 200
        assert "spreadsheetml" in resp.headers["content-type"]
        assert "attachment" in resp.headers["content-disposition"]

    def test_upload_unsupported_type(self):
        resp = self.client.post(
            "/api/upload",
            files={"file": ("test.txt", b"not image", "text/plain")},
        )
        assert resp.status_code == 415

    def test_generate_from_json(self):
        resp = self.client.post(
            "/api/generate",
            json={"headers": ["A", "B"], "data_rows": [["1", "2"]]},
        )
        assert resp.status_code == 200
        assert "spreadsheetml" in resp.headers["content-type"]

    def test_generate_empty_headers(self):
        resp = self.client.post("/api/generate", json={"headers": [], "data_rows": []})
        assert resp.status_code == 422

    def test_frontend_served(self):
        resp = self.client.get("/")
        assert resp.status_code == 200
        assert "表格解析" in resp.text
