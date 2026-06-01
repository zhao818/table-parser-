"""
Dynamic Excel rendering engine based on openpyxl.

Layout features:
- Section detection (header/body/notes)
- Freeze header row
- Auto column widths + row heights
- Print setup (landscape, fit to page)
- Merge cell support
"""

from __future__ import annotations

import io
import logging
from typing import Optional

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.page import PageMargins

from .models import TableData

logger = logging.getLogger(__name__)

THIN = Side(style="thin")
MEDIUM = Side(style="medium")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

_CJK = 2.2
_LATIN = 1.1
_MIN_W = 5.0
_MAX_W = 48.0


def generate_excel(
    table_data: TableData,
    sheet_name: str = "Sheet1",
    title: Optional[str] = None,
    auto_width: bool = True,
    first_row_as_header: bool = True,
    image_aspect_ratio: float = 0.0,  # w/h, 0 = auto-detect from data
) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    _populate(ws, table_data, sheet_name, title, auto_width, first_row_as_header, image_aspect_ratio)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def generate_excel_simple(
    headers: list[str], data_rows: list[list[str]],
    sheet_name: str = "Sheet1", title: Optional[str] = None,
) -> io.BytesIO:
    rows = [[{"text": h, "rowspan": 1, "colspan": 1} for h in headers]]
    for dr in data_rows:
        rows.append([{"text": str(c), "rowspan": 1, "colspan": 1} for c in dr])
    return generate_excel(TableData(title=title or "", rows=rows), sheet_name, title)


def generate_excel_multi_sheet(
    sheets: dict[str, TableData], title: Optional[str] = None,
) -> io.BytesIO:
    wb = Workbook()
    wb.remove(wb.active)
    for i, (sn, td) in enumerate(sheets.items()):
        ws = wb.create_sheet(title=sn[:31])
        _populate(ws, td, sn[:31], title if i == 0 else None, True, True)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def validate_table_json(data: dict) -> TableData:
    if not isinstance(data, dict):
        raise ValueError("Table JSON must be a dict")
    if "rows" not in data:
        raise ValueError("Table JSON must contain 'rows' key")
    if not isinstance(data["rows"], list):
        raise ValueError("'rows' must be a list")
    return TableData(**data)


# ---------------------------------------------------------------------------
# Internal
# ---------------------------------------------------------------------------


def _detect_sections(rows: list) -> list[tuple[int, int, int]]:
    if not rows:
        return []
    sections = []
    prev = sum(c.colspan for c in rows[0])
    start = 0
    for i, row in enumerate(rows):
        cur = sum(c.colspan for c in row)
        if prev > 0 and abs(cur - prev) > prev * 0.4:
            sections.append((start, i, prev))
            start = i
            prev = cur
    sections.append((start, len(rows), prev))
    return sections


def _populate(
    ws, table_data: TableData, sheet_name: str,
    title: Optional[str], auto_width: bool, first_row_as_header: bool,
    image_aspect_ratio: float = 0.0,
) -> None:
    ws.title = sheet_name[:31]
    display_title = title or table_data.title
    rows = table_data.rows
    if not rows:
        return

    sections = _detect_sections(rows)
    max_col = max(sum(c.colspan for c in r) for r in rows)
    cr = 1
    data_start = cr

    # Title row
    if display_title:
        c = ws.cell(row=cr, column=1, value=display_title)
        c.font = Font(bold=True, size=14)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(start_row=cr, start_column=1, end_row=cr, end_column=max_col)
        ws.row_dimensions[cr].height = 30
        cr += 1
        data_start = cr

    occupied: set[tuple[int, int]] = set()

    for sec_idx, (sec_start, sec_end, sec_cols) in enumerate(sections):
        is_header = first_row_as_header and sec_idx == 0
        is_notes = sec_idx == len(sections) - 1 and sec_idx > 0

        for r_idx in range(sec_start, sec_end):
            row_cells = rows[r_idx]
            c_idx = 1

            # Section gap
            if r_idx == sec_start and sec_idx > 0:
                cr += 1

            for cell_data in row_cells:
                while (cr, c_idx) in occupied:
                    c_idx += 1

                text, rs, cs = cell_data.text, cell_data.rowspan, cell_data.colspan
                cell = ws.cell(row=cr, column=c_idx, value=text)

                # Styling
                if is_header and r_idx == sec_start:
                    cell.font = Font(bold=True, size=10)
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                elif is_notes:
                    cell.font = Font(size=9)
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                elif c_idx == 1 and text.strip():
                    cell.font = Font(bold=True, size=10)
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                else:
                    cell.font = Font(size=10)
                    ha = "center" if len(text) < 25 else "left"
                    cell.alignment = Alignment(horizontal=ha, vertical="center", wrap_text=True)

                # Merge
                if rs > 1 or cs > 1:
                    er, ec = cr + rs - 1, c_idx + cs - 1
                    ws.merge_cells(start_row=cr, start_column=c_idx, end_row=er, end_column=ec)
                    for ri in range(cr, er + 1):
                        for ci in range(c_idx, ec + 1):
                            occupied.add((ri, ci))

                cell.border = THIN_BORDER
                c_idx += cs
            cr += 1

    last_row = cr - 1

    # Borders for all data cells
    for row in ws.iter_rows(min_row=data_start, max_row=last_row, min_col=1, max_col=max_col):
        for cell in row:
            if cell.value is not None:
                try:
                    if cell.border.left.style is None:
                        cell.border = THIN_BORDER
                except Exception:
                    pass

    # Column widths
    if auto_width:
        for col_idx in range(1, max_col + 1):
            mw = _MIN_W
            letter = get_column_letter(col_idx)
            for ri in range(data_start, last_row + 1):
                cell = ws.cell(row=ri, column=col_idx)
                if cell.value:
                    text = str(cell.value)
                    w = sum(_CJK if ord(ch) > 0x2E80 else _LATIN for ch in text)
                    mw = max(mw, min(w + 3, _MAX_W))
            ws.column_dimensions[letter].width = mw

    # Row heights
    for ri in range(data_start, last_row + 1):
        max_lines = 1
        for ci in range(1, max_col + 1):
            cell = ws.cell(row=ri, column=ci)
            if cell.value:
                text = str(cell.value)
                col_w = ws.column_dimensions[get_column_letter(ci)].width or 10
                cpl = max(1, int(col_w / 1.8))
                lines = sum(1 + max(0, (len(p) - 1) // cpl) for p in text.split("\n"))
                max_lines = max(max_lines, lines)
        if max_lines > 1:
            ws.row_dimensions[ri].height = max(20, max_lines * 15)

    # Freeze panes: only for tables with many rows (>10 data rows)
    total_data_rows = last_row - data_start + 1
    if total_data_rows > 10:
        freeze_at = data_start + 1  # freeze after header
        ws.freeze_panes = ws.cell(row=freeze_at, column=1)

    # Print settings: orientation follows image aspect ratio
    if image_aspect_ratio > 0:
        # Follow the original image's orientation
        if image_aspect_ratio >= 1.0:  # w >= h → landscape
            ws.page_setup.orientation = "landscape"
        else:
            ws.page_setup.orientation = "portrait"
    else:
        # Fallback: guess from data dimensions
        if max_col >= 8 and max_col > total_data_rows * 1.5:
            ws.page_setup.orientation = "landscape"
        else:
            ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(left=0.3, right=0.3, top=0.4, bottom=0.4)
